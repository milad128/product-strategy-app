"""Parse monthly lifecycle count files (Excel / CSV)."""

from __future__ import annotations

import csv
import io
import re
from typing import Any

STAGE_LABEL_TO_ID: dict[str, str] = {
    "abandoned": "abandoned",
    "abandoned_user": "abandoned",
    "active_customer": "activeCustomer",
    "active_customer_6m": "activeCustomer",
    "active_customer_1m": "custom-1783425377379",
    "dead_credit_holder": "deadCreditHolder",
    "dead_closed_customer": "creditClosed",
    "dead_customer": "creditClosed",
    "dormant_customer": "dormantCustomer",
    "dormant_credit_holder": "dormantCustomer",
    "fresh_credit_holder": "freshCreditHolder",
    "rejected": "rejected",
    "rejected_user": "rejected",
    "soft_churn_customer": "softChurned",
    "unactivated_credit_holder": "unActivatedCreditHolder",
    "applicant": "applicant",
    "applicant_user": "applicant",
}

BUILTIN_STAGE_IDS = set(STAGE_LABEL_TO_ID.values()) | {"blackList"}

MONTH_PATTERN = re.compile(r"^\d{6}$")


def normalize_stage_label(label: str) -> str:
    return re.sub(r"\s+", "_", label.strip().lower())


def normalize_month(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()
    if not text or text.lower() in {"none", "nan"}:
        return None
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    if MONTH_PATTERN.match(text):
        return text
    return None


def parse_count(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value if value >= 0 else None
    if isinstance(value, float):
        if not value.is_integer() or value < 0:
            return None
        return int(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        number = int(text)
    except ValueError:
        return None
    return number if number >= 0 else None


def _grid_from_rows(rows: list[list[Any]]) -> tuple[dict[str, dict[str, int]], list[str]]:
    if len(rows) < 2:
        raise ValueError("File must have a header row and at least one stage row.")

    header = rows[0]
    months: list[tuple[int, str]] = []
    for col_idx in range(1, len(header)):
        month = normalize_month(header[col_idx])
        if month:
            months.append((col_idx, month))

    if not months:
        raise ValueError("No valid Jalali month columns found in row 1 (expected YYYYMM).")

    by_month: dict[str, dict[str, int | None]] = {m: {} for _, m in months}
    warnings: list[str] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        if not row or not str(row[0]).strip():
            continue
        label = normalize_stage_label(str(row[0]))
        stage_id = STAGE_LABEL_TO_ID.get(label)
        if not stage_id:
            warnings.append(f"Row {row_idx}: unknown stage {row[0]!r} (skipped)")
            continue
        for col_idx, month in months:
            raw = row[col_idx] if col_idx < len(row) else None
            count = parse_count(raw)
            if raw not in (None, "") and count is None:
                raise ValueError(
                    f"Invalid count at row {row_idx}, month {month}: {raw!r} "
                    "(must be a non-negative integer)."
                )
            by_month[month][stage_id] = count if count is not None else 0

    result: dict[str, dict[str, int]] = {}
    for month, stage_counts in by_month.items():
        counts: dict[str, int] = {sid: 0 for sid in BUILTIN_STAGE_IDS}
        counts.update({k: v for k, v in stage_counts.items() if v is not None})
        counts["blackList"] = counts.get("blackList", 0)
        result[month] = counts

    return result, warnings


def parse_csv_bytes(data: bytes) -> tuple[dict[str, dict[str, int]], list[str]]:
    text = data.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = [list(row) for row in reader]
    return _grid_from_rows(rows)


def parse_xlsx_bytes(data: bytes) -> tuple[dict[str, dict[str, int]], list[str]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ValueError("Excel import requires openpyxl. Install project requirements.") from exc

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows: list[list[Any]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(list(row))
    wb.close()
    return _grid_from_rows(rows)


def parse_monthly_file(filename: str, data: bytes) -> tuple[dict[str, dict[str, int]], list[str]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return parse_csv_bytes(data)
    if lower.endswith(".xlsx"):
        return parse_xlsx_bytes(data)
    raise ValueError("Unsupported file type. Upload .xlsx or .csv.")
