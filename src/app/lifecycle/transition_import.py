"""Parse avg_transition_probs_<month>.xlsx files (10x10 stage transition matrix)."""

from __future__ import annotations

import io
import re
from typing import Any

from src.app.lifecycle.monthly_import import STAGE_LABEL_TO_ID, normalize_stage_label

FILENAME_MONTH = re.compile(r"(\d{6})")


def month_from_filename(filename: str) -> str | None:
    match = FILENAME_MONTH.search(filename)
    return match.group(1) if match else None


def parse_probability(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        n = float(value)
    except (TypeError, ValueError):
        return None
    if n < 0:
        return None
    return min(1.0, n)


def parse_transition_probs_xlsx(
    filename: str, data: bytes
) -> tuple[str, dict[str, dict[str, float]]]:
    """Return (month, {from_stage_id: {to_stage_id: probability}}), skipping self-loops and unmapped segments."""
    try:
        import openpyxl
    except ImportError as exc:
        raise ValueError("Excel import requires openpyxl. Install project requirements.") from exc

    month = month_from_filename(filename)
    if not month:
        raise ValueError(
            "Could not find a Jalali month (YYYYMM) in the filename, e.g. avg_transition_probs_140401.xlsx"
        )

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    wb.close()

    if len(rows) < 2:
        raise ValueError("File must have a header row and at least one stage row.")

    header = rows[0]
    columns: list[tuple[int, str]] = []
    for col_idx in range(1, len(header)):
        label = header[col_idx]
        if label is None:
            continue
        stage_id = STAGE_LABEL_TO_ID.get(normalize_stage_label(str(label)))
        if stage_id:
            columns.append((col_idx, stage_id))

    if not columns:
        raise ValueError("No recognized stage columns found in row 1.")

    matrix: dict[str, dict[str, float]] = {}
    for row in rows[1:]:
        if not row or row[0] is None or not str(row[0]).strip():
            continue
        from_id = STAGE_LABEL_TO_ID.get(normalize_stage_label(str(row[0])))
        if not from_id:
            continue
        for col_idx, to_id in columns:
            if to_id == from_id:
                continue  # self-loop — not a connector
            raw = row[col_idx] if col_idx < len(row) else None
            prob = parse_probability(raw)
            if prob is None or prob <= 0:
                continue
            matrix.setdefault(from_id, {})[to_id] = prob

    return month, matrix
